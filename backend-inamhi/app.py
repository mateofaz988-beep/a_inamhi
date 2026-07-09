from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
import mysql.connector
import json
import os
import glob
import copy
import inspect
import jwt
import datetime
from io import BytesIO
from datetime import date, datetime, timezone, timedelta
from decimal import Decimal
from services.pdf_converter import PdfConverter
from services.firma_service import FirmaService
from services.documento_service import DocumentoService
from services.certificado_service import CertificadoService
import config_firmas
from config import DB_CONFIG, FLASK_PORT, FLASK_DEBUG, MAX_CERTIFICADO_BYTES, JWT_SECRET_KEY

try:
    import openpyxl
    OPENPYXL_DISPONIBLE = True
except ImportError:
    OPENPYXL_DISPONIBLE = False

app = Flask(__name__)
CORS(app, resources={r"/api/*": {"origins": "*"}}, supports_credentials=True)

# Clave secreta para firmar los JWT — viene de config.py (JWT_SECRET_KEY en .env).
# Antes se leía os.environ.get('SECRET_KEY', ...), una variable que .env nunca
# definió (el archivo usa JWT_SECRET_KEY), así que siempre se usaba el valor
# de respaldo hardcodeado aquí mismo, visible en el repositorio.
SECRET_KEY = JWT_SECRET_KEY

db_config = DB_CONFIG

# =========================
# 🔌 CONEXIÓN A MYSQL
# =========================
def get_connection():
    conexion = mysql.connector.connect(**db_config)
    cursor = conexion.cursor()
    cursor.execute("SET time_zone = '-05:00'")
    cursor.close()
    return conexion

# =========================
# 🔄 SERIALIZADOR JSON
# =========================
def json_serializer(obj):
    if isinstance(obj, (datetime, date)):
        return obj.isoformat()
    if isinstance(obj, Decimal):
        return float(obj)
    return str(obj)

# =========================
# 🕒 FECHA/HORA ECUADOR
# =========================
def fecha_ecuador():
    tz_ec = timezone(timedelta(hours=-5))
    return datetime.now(tz_ec).strftime('%Y-%m-%d %H:%M:%S')


# =========================
# 📄 UTILIDADES PDF PARA FIRMAS
# =========================
def obtener_total_paginas_pdf(ruta_pdf):
    """Devuelve el total de páginas del PDF y valida que sea firmable."""
    if not ruta_pdf:
        raise ValueError("No se recibió la ruta del PDF.")

    ruta_pdf = os.path.abspath(ruta_pdf)
    if not os.path.exists(ruta_pdf):
        raise FileNotFoundError(f"No existe el PDF: {ruta_pdf}")

    try:
        try:
            from pypdf import PdfReader
        except ImportError:
            from pypdf import PdfReader

        lector = PdfReader(ruta_pdf)
        total = len(lector.pages)
    except Exception as error:
        raise ValueError(f"No se pudo leer el PDF generado: {error}") from error

    if total <= 0:
        raise ValueError("El PDF generado no contiene páginas.")

    return total


def normalizar_indice_pagina(pagina_configurada, total_paginas):
    """
    Convierte la página configurada en un índice válido para pyHanko.

    pyHanko usa índices desde cero:
      página visual 1 -> índice 0
      página visual 2 -> índice 1
    """
    if total_paginas <= 0:
        raise ValueError("El PDF no contiene páginas.")

    try:
        pagina = int(pagina_configurada)
    except (TypeError, ValueError):
        pagina = 0

    if pagina < 0:
        pagina = total_paginas + pagina

    if pagina < 0:
        pagina = 0
    elif pagina >= total_paginas:
        pagina = total_paginas - 1

    return pagina


def normalizar_posicion_firma(posicion, total_paginas, pagina_preferida=0):
    """
    Conserva las coordenadas originales y corrige únicamente el índice
    de página cuando la configuración incluye página/on_page/page.
    """
    pagina_valida = normalizar_indice_pagina(pagina_preferida, total_paginas)
    posicion_normalizada = copy.deepcopy(posicion)

    if isinstance(posicion_normalizada, dict):
        claves_pagina = (
            'pagina', 'page', 'on_page', 'page_index', 'indice_pagina'
        )
        clave_encontrada = next(
            (clave for clave in claves_pagina if clave in posicion_normalizada),
            None
        )

        if clave_encontrada:
            posicion_normalizada[clave_encontrada] = pagina_valida
        else:
            # No elimina ni altera x1/y1/x2/y2 o box.
            posicion_normalizada['pagina'] = pagina_valida

    elif isinstance(posicion_normalizada, (list, tuple)):
        # Algunas implementaciones usan (pagina, x1, y1, x2, y2).
        if len(posicion_normalizada) == 5:
            valores = list(posicion_normalizada)
            valores[0] = pagina_valida
            posicion_normalizada = (
                tuple(valores)
                if isinstance(posicion_normalizada, tuple)
                else valores
            )

    return posicion_normalizada, pagina_valida


def ejecutar_firma_pdf(
    ruta_entrada,
    ruta_salida,
    cert_info,
    seccion,
    nombre_firmante,
    cargo_firmante,
    posicion,
    pagina
):
    """
    Invoca FirmaService.firmar_pdf y transmite la página si el servicio
    declara un parámetro compatible. Mantiene compatibilidad con la firma
    anterior que recibía solamente la posición.
    """
    parametros = inspect.signature(FirmaService.firmar_pdf).parameters
    kwargs = {}

    for nombre_parametro in (
        'pagina', 'page', 'on_page', 'page_index', 'indice_pagina'
    ):
        if nombre_parametro in parametros:
            kwargs[nombre_parametro] = pagina
            break

    return FirmaService.firmar_pdf(
        ruta_entrada,
        ruta_salida,
        cert_info,
        seccion,
        nombre_firmante,
        cargo_firmante,
        posicion,
        **kwargs
    )

# =========================
# 🔐 DECODIFICAR TOKEN
# =========================
def decodificar_token(auth_header):
    if not auth_header or not auth_header.startswith('Bearer '):
        return None
    token = auth_header.split(' ')[1]
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=['HS256'])
        return payload
    except jwt.ExpiredSignatureError:
        return None
    except jwt.InvalidTokenError:
        return None

# =========================
# 🔐 OBTENER USUARIO DEL TOKEN
# =========================
def obtener_usuario():
    auth_header = request.headers.get('Authorization')
    payload = decodificar_token(auth_header)
    if payload and 'usuario' in payload:
        return payload['usuario']
    return 'desconocido'

# =========================
# 🌐 OBTENER IP
# =========================
def obtener_ip():
    forwarded = request.headers.get('X-Forwarded-For')
    if forwarded:
        return forwarded.split(',')[0].strip()
    return request.remote_addr

# =========================
# 🔐 VALIDAR ADMIN
# =========================
import traceback
from werkzeug.exceptions import HTTPException

@app.errorhandler(HTTPException)
def handle_http_exception(e):
    """Retorna el código HTTP correcto para 404, 403, etc. sin loggear como error 500."""
    return jsonify({"error": e.description}), e.code

@app.errorhandler(Exception)
def handle_exception(e):
    """Solo captura excepciones reales del servidor (no HTTPException)."""
    with open('error_500.log', 'a') as f:
        f.write(traceback.format_exc())
    return jsonify({"error": str(e)}), 500

def es_admin():
    auth_header = request.headers.get('Authorization')
    payload = decodificar_token(auth_header)
    if payload and 'rol' in payload:
        return payload['rol'] == 'admin'
    return False

# =========================
# 🧾 NORMALIZAR GÉNERO
# =========================
def normalizar_genero(genero):
    if genero is None:
        return None

    genero = str(genero).strip().lower()

    if genero in ['m', 'masculino']:
        return 'Masculino'

    if genero in ['f', 'femenino']:
        return 'Femenino'

    return genero.capitalize()

# =========================
# 🧾 REGISTRAR AUDITORÍA
# =========================
def registrar_auditoria(usuario, accion, tabla, registro_id=None, antes=None, despues=None, detalle=None):
    conexion = None
    cursor = None

    try:
        conexion = get_connection()
        cursor = conexion.cursor()

        query = """
            INSERT INTO auditoria
            (usuario, accion, tabla_afectada, registro_id, datos_anteriores, datos_nuevos, detalle, ip_usuario, fecha)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """

        datos_anteriores = json.dumps(antes, ensure_ascii=False, default=json_serializer) if antes is not None else None
        datos_nuevos = json.dumps(despues, ensure_ascii=False, default=json_serializer) if despues is not None else None

        cursor.execute(query, (
            usuario,
            accion,
            tabla,
            registro_id,
            datos_anteriores,
            datos_nuevos,
            detalle,
            obtener_ip(),
            fecha_ecuador()
        ))

        conexion.commit()

    except Exception as e:
        print("ERROR REGISTRANDO AUDITORIA:", str(e))

    finally:
        if cursor:
            cursor.close()
        if conexion:
            conexion.close()

# =========================
# 🔐 LOGIN
# =========================
@app.route('/api/auth/login', methods=['POST'])
def login():
    data = request.get_json()

    conexion = None
    cursor = None

    try:
        conexion = get_connection()
        cursor = conexion.cursor(dictionary=True)

        query = """
            SELECT usuario, rol
            FROM usuarios
            WHERE usuario = %s AND password = %s
        """
        cursor.execute(query, (data.get('user'), data.get('pass')))
        usuario = cursor.fetchone()

        if usuario:
            registrar_auditoria(
                usuario=usuario['usuario'],
                accion='LOGIN',
                tabla='usuarios',
                registro_id=None,
                antes=None,
                despues={"usuario": usuario['usuario'], "rol": usuario['rol']},
                detalle='Inicio de sesión exitoso'
            )

            try:
                exp_date = datetime.now(timezone.utc) + timedelta(hours=12)
                jwt_token = jwt.encode({
                    'usuario': usuario['usuario'],
                    'rol': usuario['rol'],
                    'exp': exp_date
                }, SECRET_KEY, algorithm='HS256')
            except Exception as token_err:
                print(f"ERROR GENERANDO TOKEN JWT: {token_err}")
                return jsonify({"error": "Error interno generando credenciales"}), 500

            return jsonify({
                "token": jwt_token,
                "role": usuario['rol'],
                "usuario": usuario['usuario']
            }), 200

        return jsonify({"error": "No autorizado"}), 401

    except Exception as e:
        return jsonify({"error": str(e)}), 500

    finally:
        if cursor:
            cursor.close()
        if conexion:
            conexion.close()

# =========================
# 📥 GET PERSONAL
# =========================
@app.route('/api/personal', methods=['GET'])
def obtener_personal():
    token = request.headers.get('Authorization', '')
    if not decodificar_token(token):
        return jsonify({"error": "No autorizado"}), 403

    conexion = None
    cursor = None

    try:
        conexion = get_connection()
        cursor = conexion.cursor(dictionary=True)

        cursor.execute("SELECT * FROM personal")
        resultados = cursor.fetchall()

        return jsonify(resultados), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500

    finally:
        if cursor:
            cursor.close()
        if conexion:
            conexion.close()

# =========================
# 🔎 GET PERSONAL POR CÉDULA
# =========================
@app.route('/api/personal/cedula/<cedula>', methods=['GET'])
def obtener_personal_por_cedula(cedula):
    if not es_admin():
        return jsonify({"error": "No autorizado"}), 403

    conexion = None
    cursor = None

    try:
        conexion = get_connection()
        cursor = conexion.cursor(dictionary=True)

        query = """
            SELECT *
            FROM personal
            WHERE cedula = %s
            LIMIT 1
        """
        cursor.execute(query, (cedula,))
        resultado = cursor.fetchone()

        if not resultado:
            return jsonify({"error": "No se encontró información para esa cédula"}), 404

        return jsonify(resultado), 200

    except Exception as e:
        print("ERROR CONSULTANDO CEDULA:", str(e))
        return jsonify({"error": str(e)}), 500

    finally:
        if cursor:
            cursor.close()
        if conexion:
            conexion.close()

# =========================
# ➕ CREATE PERSONAL
# =========================
@app.route('/api/personal', methods=['POST'])
def crear_personal():
    if not es_admin():
        return jsonify({"error": "No autorizado"}), 403

    data = request.get_json()
    usuario = obtener_usuario()

    conexion = None
    cursor = None

    try:
        conexion = get_connection()
        cursor = conexion.cursor()

        genero = normalizar_genero(data.get('genero'))

        query = """
            INSERT INTO personal (
                nro, cedula, nombres, modalidad, cargo, rmu, unidad,
                fecha_ingreso, fecha_nacimiento, direccion, email_inst,
                telefono, genero, instruccion, profesion, vulnerable,
                tipo_discapacidad, porcentaje_disc, etnia, observaciones
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """

        valores = (
            data.get('nro'),
            data.get('cedula'),
            data.get('nombres'),
            data.get('modalidad'),
            data.get('cargo'),
            data.get('rmu'),
            data.get('unidad'),
            data.get('fecha_ingreso'),
            data.get('fecha_nacimiento'),
            data.get('direccion'),
            data.get('email_inst'),
            data.get('telefono'),
            genero,
            data.get('instruccion'),
            data.get('profesion'),
            data.get('vulnerable'),
            data.get('tipo_discapacidad'),
            data.get('porcentaje_disc'),
            data.get('etnia'),
            data.get('observaciones')
        )

        cursor.execute(query, valores)
        conexion.commit()

        nuevo_id = cursor.lastrowid

        registrar_auditoria(
            usuario=usuario,
            accion='CREATE',
            tabla='personal',
            registro_id=nuevo_id,
            antes=None,
            despues=data,
            detalle='Creación de funcionario'
        )

        return jsonify({"message": "Creado correctamente"}), 201

    except Exception as e:
        print("ERROR CREATE PERSONAL:", str(e))
        return jsonify({"error": str(e)}), 500

    finally:
        if cursor:
            cursor.close()
        if conexion:
            conexion.close()

# =========================
# ✏️ UPDATE PERSONAL
# =========================
@app.route('/api/personal/<int:id>', methods=['PUT'])
def actualizar_personal(id):
    if not es_admin():
        return jsonify({"error": "No autorizado"}), 403

    data = request.get_json()
    usuario = obtener_usuario()

    conexion = None
    cursor = None

    try:
        conexion = get_connection()
        cursor = conexion.cursor(dictionary=True)

        cursor.execute("SELECT * FROM personal WHERE id = %s", (id,))
        antes = cursor.fetchone()

        if not antes:
            return jsonify({"error": "Registro no encontrado"}), 404

        genero = normalizar_genero(data.get('genero'))

        query = """
            UPDATE personal
            SET
                nro=%s,
                cedula=%s,
                nombres=%s,
                modalidad=%s,
                cargo=%s,
                rmu=%s,
                unidad=%s,
                fecha_ingreso=%s,
                fecha_nacimiento=%s,
                direccion=%s,
                email_inst=%s,
                telefono=%s,
                genero=%s,
                instruccion=%s,
                profesion=%s,
                vulnerable=%s,
                tipo_discapacidad=%s,
                porcentaje_disc=%s,
                etnia=%s,
                rol=%s,
                observaciones=%s
            WHERE id=%s
        """

        cursor.execute(query, (
            data.get('nro'),
            data.get('cedula'),
            data.get('nombres'),
            data.get('modalidad'),
            data.get('cargo'),
            data.get('rmu'),
            data.get('unidad'),
            data.get('fecha_ingreso'),
            data.get('fecha_nacimiento'),
            data.get('direccion'),
            data.get('email_inst'),
            data.get('telefono'),
            genero,
            data.get('instruccion'),
            data.get('profesion'),
            data.get('vulnerable'),
            data.get('tipo_discapacidad'),
            data.get('porcentaje_disc'),
            data.get('etnia'),
            data.get('rol'),
            data.get('observaciones'),
            id
        ))

        conexion.commit()

        registrar_auditoria(
            usuario=usuario,
            accion='UPDATE',
            tabla='personal',
            registro_id=id,
            antes=antes,
            despues=data,
            detalle='Actualización de funcionario'
        )

        return jsonify({"message": "Actualizado correctamente"}), 200

    except Exception as e:
        print("ERROR UPDATE PERSONAL:", str(e))
        return jsonify({"error": str(e)}), 500

    finally:
        if cursor:
            cursor.close()
        if conexion:
            conexion.close()

# =========================
# ❌ DELETE PERSONAL
# =========================
@app.route('/api/personal/<int:id>', methods=['DELETE'])
def eliminar_personal(id):
    if not es_admin():
        return jsonify({"error": "No autorizado"}), 403

    usuario = obtener_usuario()

    conexion = None
    cursor = None

    try:
        conexion = get_connection()
        cursor = conexion.cursor(dictionary=True)

        cursor.execute("SELECT * FROM personal WHERE id = %s", (id,))
        antes = cursor.fetchone()

        cursor.execute("DELETE FROM personal WHERE id = %s", (id,))
        conexion.commit()

        registrar_auditoria(
            usuario=usuario,
            accion='DELETE',
            tabla='personal',
            registro_id=id,
            antes=antes,
            despues=None,
            detalle='Eliminación de funcionario'
        )

        return jsonify({"message": "Eliminado"}), 200

    except Exception as e:
        print("ERROR DELETE PERSONAL:", str(e))
        return jsonify({"error": str(e)}), 500

    finally:
        if cursor:
            cursor.close()
        if conexion:
            conexion.close()

# =========================
# 📊 GET AUDITORIA
# =========================
@app.route('/api/auditoria', methods=['GET'])
def obtener_auditoria():
    if not es_admin():
        return jsonify({"error": "No autorizado"}), 403

    conexion = None
    cursor = None

    try:
        conexion = get_connection()
        cursor = conexion.cursor(dictionary=True)

        query = """
            SELECT
                id,
                usuario,
                accion,
                tabla_afectada,
                registro_id,
                datos_anteriores,
                datos_nuevos,
                detalle,
                ip_usuario,
                DATE_FORMAT(fecha, '%Y-%m-%d %H:%i:%s') AS fecha
            FROM auditoria
            ORDER BY fecha DESC, id DESC
        """

        cursor.execute(query)
        resultados = cursor.fetchall()

        return jsonify(resultados), 200

    except Exception as e:
        print("ERROR AUDITORIA:", str(e))
        return jsonify({"error": str(e)}), 500

    finally:
        if cursor:
            cursor.close()
        if conexion:
            conexion.close()
# =========================
# 📋 GET USUARIOS
# =========================
@app.route('/api/usuarios', methods=['GET'])
def obtener_usuarios():
    if not es_admin():
        return jsonify({"error": "No autorizado"}), 403

    conexion = None
    cursor = None

    try:
        conexion = get_connection()
        cursor = conexion.cursor(dictionary=True)

        cursor.execute("""
            SELECT id, usuario, rol
            FROM usuarios
            ORDER BY id ASC
        """)

        resultados = cursor.fetchall()

        return jsonify(resultados), 200

    except Exception as e:
        print("ERROR OBTENIENDO USUARIOS:", str(e))
        return jsonify({"error": str(e)}), 500

    finally:
        if cursor:
            cursor.close()
        if conexion:
            conexion.close()


# =========================
# 👤 CREATE USUARIO
# =========================
@app.route('/api/usuarios', methods=['POST'])
def crear_usuario():
    if not es_admin():
        return jsonify({"error": "No autorizado"}), 403

    data = request.get_json()
    usuario_admin = obtener_usuario()

    usuario_nuevo = (data.get('usuario') or '').strip()
    password_nueva = (data.get('password') or '').strip()
    rol_nuevo = (data.get('rol') or '').strip().lower()

    # VALIDACIONES
    if not usuario_nuevo:
        return jsonify({"error": "El usuario es obligatorio"}), 400

    if len(usuario_nuevo) < 4:
        return jsonify({"error": "El usuario debe tener al menos 4 caracteres"}), 400

    if len(usuario_nuevo) > 30:
        return jsonify({"error": "El usuario no debe superar los 30 caracteres"}), 400

    if not password_nueva:
        return jsonify({"error": "La contraseña es obligatoria"}), 400

    if len(password_nueva) < 4:
        return jsonify({"error": "La contraseña debe tener al menos 4 caracteres"}), 400

    if rol_nuevo not in ['admin', 'visitante']:
        return jsonify({"error": "Rol inválido"}), 400

    conexion = None
    cursor = None

    try:
        conexion = get_connection()
        cursor = conexion.cursor(dictionary=True)

        # VALIDAR DUPLICADO
        cursor.execute(
            "SELECT id FROM usuarios WHERE usuario = %s",
            (usuario_nuevo,)
        )
        existe = cursor.fetchone()

        if existe:
            return jsonify({"error": "El usuario ya existe"}), 409

        cursor.close()
        cursor = conexion.cursor()

        query = """
            INSERT INTO usuarios (usuario, password, rol)
            VALUES (%s, %s, %s)
        """

        cursor.execute(query, (usuario_nuevo, password_nueva, rol_nuevo))
        conexion.commit()

        nuevo_id = cursor.lastrowid

        # AUDITORÍA
        registrar_auditoria(
            usuario=usuario_admin,
            accion='CREATE',
            tabla='usuarios',
            registro_id=nuevo_id,
            antes=None,
            despues={
                "usuario": usuario_nuevo,
                "rol": rol_nuevo
            },
            detalle='Creación de usuario'
        )

        return jsonify({"message": "Usuario creado correctamente"}), 201

    except Exception as e:
        print("ERROR CREANDO USUARIO:", str(e))
        return jsonify({"error": str(e)}), 500

    finally:
        if cursor:
            cursor.close()
        if conexion:
            conexion.close()


# =========================
# ✏️ UPDATE ROL USUARIO
# =========================
@app.route('/api/usuarios/<int:id>', methods=['PUT'])
def actualizar_usuario(id):
    if not es_admin():
        return jsonify({"error": "No autorizado"}), 403

    data = request.get_json()
    usuario_admin = obtener_usuario()

    nuevo_rol = (data.get('rol') or '').strip().lower()

    if nuevo_rol not in ['admin', 'visitante']:
        return jsonify({"error": "Rol inválido"}), 400

    conexion = None
    cursor = None

    try:
        conexion = get_connection()
        cursor = conexion.cursor(dictionary=True)

        # OBTENER ANTES
        cursor.execute("SELECT * FROM usuarios WHERE id = %s", (id,))
        antes = cursor.fetchone()

        if not antes:
            return jsonify({"error": "Usuario no encontrado"}), 404

        # 🔥 NO TE PUEDES QUITAR ADMIN A TI MISMO
        if antes['usuario'] == usuario_admin and nuevo_rol != 'admin':
            return jsonify({"error": "No puedes quitarte el rol admin"}), 400

        cursor.close()
        cursor = conexion.cursor()

        cursor.execute(
            "UPDATE usuarios SET rol = %s WHERE id = %s",
            (nuevo_rol, id)
        )
        conexion.commit()

        # AUDITORÍA
        registrar_auditoria(
            usuario=usuario_admin,
            accion='UPDATE',
            tabla='usuarios',
            registro_id=id,
            antes=antes,
            despues={
                "rol": nuevo_rol
            },
            detalle='Actualización de rol'
        )

        return jsonify({"message": "Rol actualizado correctamente"}), 200

    except Exception as e:
        print("ERROR ACTUALIZANDO USUARIO:", str(e))
        return jsonify({"error": str(e)}), 500

    finally:
        if cursor:
            cursor.close()
        if conexion:
            conexion.close()


# =========================
# ❌ DELETE USUARIO
# =========================
@app.route('/api/usuarios/<int:id>', methods=['DELETE'])
def eliminar_usuario(id):
    if not es_admin():
        return jsonify({"error": "No autorizado"}), 403

    usuario_admin = obtener_usuario()

    conexion = None
    cursor = None

    try:
        conexion = get_connection()
        cursor = conexion.cursor(dictionary=True)

        cursor.execute("SELECT * FROM usuarios WHERE id = %s", (id,))
        antes = cursor.fetchone()

        if not antes:
            return jsonify({"error": "Usuario no encontrado"}), 404

        # 🔥 NO PUEDES ELIMINARTE A TI MISMO
        if antes['usuario'] == usuario_admin:
            return jsonify({"error": "No puedes eliminar tu propio usuario"}), 400

        cursor.close()
        cursor = conexion.cursor()

        cursor.execute("DELETE FROM usuarios WHERE id = %s", (id,))
        conexion.commit()

        registrar_auditoria(
            usuario=usuario_admin,
            accion='DELETE',
            tabla='usuarios',
            registro_id=id,
            antes=antes,
            despues=None,
            detalle='Eliminación de usuario'
        )

        return jsonify({"message": "Usuario eliminado correctamente"}), 200

    except Exception as e:
        print("ERROR ELIMINANDO USUARIO:", str(e))
        return jsonify({"error": str(e)}), 500

    finally:
        if cursor:
            cursor.close()
        if conexion:
            conexion.close()
            # =========================
# 🚪 DESVINCULAR PERSONAL
# =========================
@app.route('/api/personal/<int:id>/desvincular', methods=['POST'])
def desvincular_personal(id):
    if not es_admin():
        return jsonify({"error": "No autorizado"}), 403

    data = request.get_json() or {}
    motivo_salida = (data.get('motivo_salida') or '').strip()
    usuario = obtener_usuario()

    if not motivo_salida or len(motivo_salida) < 5:
        return jsonify({"error": "Debe ingresar un motivo de salida válido"}), 400

    conexion = None
    cursor = None

    try:
        conexion = get_connection()
        cursor = conexion.cursor(dictionary=True)

        cursor.execute("SELECT * FROM personal WHERE id = %s", (id,))
        emp = cursor.fetchone()

        if not emp:
            return jsonify({"error": "Funcionario no encontrado"}), 404

        cursor.execute("SELECT id FROM personal_pasivo WHERE cedula = %s", (emp.get('cedula'),))
        existente = cursor.fetchone()

        if existente:
            return jsonify({"error": "Este funcionario ya consta como desvinculado"}), 409

        query_insert = """
            INSERT INTO personal_pasivo (
                id_personal, nro, cedula, nombres, modalidad, cargo, rmu, unidad,
                fecha_ingreso, fecha_nacimiento, direccion, email_inst, telefono,
                genero, instruccion, profesion, vulnerable, tipo_discapacidad,
                porcentaje_disc, etnia, rol, observaciones,
                fecha_salida, motivo_salida, usuario_responsable
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """

        cursor.execute(query_insert, (
            emp.get('id'),
            emp.get('nro'),
            emp.get('cedula'),
            emp.get('nombres'),
            emp.get('modalidad'),
            emp.get('cargo'),
            emp.get('rmu'),
            emp.get('unidad'),
            emp.get('fecha_ingreso'),
            emp.get('fecha_nacimiento'),
            emp.get('direccion'),
            emp.get('email_inst'),
            emp.get('telefono'),
            emp.get('genero'),
            emp.get('instruccion'),
            emp.get('profesion'),
            emp.get('vulnerable'),
            emp.get('tipo_discapacidad'),
            emp.get('porcentaje_disc'),
            emp.get('etnia'),
            emp.get('rol'),
            emp.get('observaciones'),
            fecha_ecuador(),
            motivo_salida,
            usuario
        ))

        cursor.execute("DELETE FROM personal WHERE id = %s", (id,))
        conexion.commit()

        registrar_auditoria(
            usuario=usuario,
            accion='DELETE',
            tabla='personal',
            registro_id=id,
            antes=emp,
            despues={
                "estado": "DESVINCULADO",
                "motivo_salida": motivo_salida
            },
            detalle='Funcionario desvinculado y movido a personal_pasivo'
        )

        return jsonify({"message": "Funcionario desvinculado correctamente"}), 200

    except Exception as e:
        if conexion:
            conexion.rollback()
        print("ERROR DESVINCULANDO PERSONAL:", str(e))
        return jsonify({"error": str(e)}), 500

    finally:
        if cursor:
            cursor.close()
        if conexion:
            conexion.close()
# =========================
# 🔄 REACTIVAR PERSONAL PASIVO
# =========================
@app.route('/api/personal/pasivo/<int:id>/reactivar', methods=['POST'])
def reactivar_personal_pasivo(id):
    if not es_admin():
        return jsonify({"error": "No autorizado"}), 403

    usuario = obtener_usuario()
    conexion = None
    cursor = None

    try:
        conexion = get_connection()
        cursor = conexion.cursor(dictionary=True)

        cursor.execute("SELECT * FROM personal_pasivo WHERE id = %s", (id,))
        emp = cursor.fetchone()

        if not emp:
            return jsonify({"error": "Funcionario pasivo no encontrado"}), 404

        cursor.execute("SELECT id FROM personal WHERE cedula = %s", (emp.get('cedula'),))
        activo = cursor.fetchone()

        if activo:
            return jsonify({"error": "Ya existe un funcionario activo con esa cédula"}), 409

        query_insert = """
            INSERT INTO personal (
                nro, cedula, nombres, modalidad, cargo, rmu, unidad,
                fecha_ingreso, fecha_nacimiento, direccion, email_inst,
                telefono, genero, instruccion, profesion, vulnerable,
                tipo_discapacidad, porcentaje_disc, etnia, rol, observaciones
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """

        cursor.execute(query_insert, (
            emp.get('nro'),
            emp.get('cedula'),
            emp.get('nombres'),
            emp.get('modalidad'),
            emp.get('cargo'),
            emp.get('rmu'),
            emp.get('unidad'),
            emp.get('fecha_ingreso'),
            emp.get('fecha_nacimiento'),
            emp.get('direccion'),
            emp.get('email_inst'),
            emp.get('telefono'),
            emp.get('genero'),
            emp.get('instruccion'),
            emp.get('profesion'),
            emp.get('vulnerable'),
            emp.get('tipo_discapacidad'),
            emp.get('porcentaje_disc'),
            emp.get('etnia'),
            emp.get('rol'),
            emp.get('observaciones')
        ))

        nuevo_id = cursor.lastrowid

        cursor.execute("DELETE FROM personal_pasivo WHERE id = %s", (id,))
        conexion.commit()

        registrar_auditoria(
            usuario=usuario,
            accion='CREATE',
            tabla='personal',
            registro_id=nuevo_id,
            antes={
                "estado": "PASIVO",
                "datos": emp
            },
            despues={
                "estado": "REACTIVADO",
                "nuevo_id": nuevo_id
            },
            detalle='Funcionario reactivado desde personal_pasivo'
        )

        return jsonify({"message": "Funcionario reactivado correctamente"}), 200

    except Exception as e:
        if conexion:
            conexion.rollback()
        print("ERROR REACTIVANDO PERSONAL:", str(e))
        return jsonify({"error": str(e)}), 500

    finally:
        if cursor:
            cursor.close()
        if conexion:
            conexion.close()
# =========================
# 📋 GET PERSONAL PASIVO
# =========================
@app.route('/api/personal/pasivo', methods=['GET'])
def obtener_personal_pasivo():
    if not es_admin():
        return jsonify({"error": "No autorizado"}), 403

    conexion = None
    cursor = None

    try:
        conexion = get_connection()
        cursor = conexion.cursor(dictionary=True)

        cursor.execute("""
            SELECT
                id,
                id_personal,
                nro,
                cedula,
                nombres,
                modalidad,
                cargo,
                rmu,
                unidad,
                fecha_ingreso,
                fecha_nacimiento,
                direccion,
                email_inst,
                telefono,
                genero,
                instruccion,
                profesion,
                vulnerable,
                tipo_discapacidad,
                porcentaje_disc,
                etnia,
                rol,
                observaciones,
                DATE_FORMAT(fecha_salida, '%Y-%m-%d %H:%i:%s') AS fecha_salida,
                motivo_salida,
                usuario_responsable,
                DATE_FORMAT(created_at, '%Y-%m-%d %H:%i:%s') AS created_at
            FROM personal_pasivo
            ORDER BY fecha_salida DESC, id DESC
        """)

        resultados = cursor.fetchall()

        return jsonify(resultados), 200

    except Exception as e:
        print("ERROR OBTENIENDO PERSONAL PASIVO:", str(e))
        return jsonify({"error": str(e)}), 500

    finally:
        if cursor:
            cursor.close()
        if conexion:
            conexion.close()
# =========================
# ⚖️ CATÁLOGO DE BASES LEGALES POR TIPO DE ACCIÓN
# =========================
@app.route('/api/base-legal', methods=['GET'])
def get_base_legal():
    token = request.headers.get('Authorization', '')
    if not decodificar_token(token):
        return jsonify({"error": "No autorizado"}), 403

    try:
        # Conexión a la base de datos (asegúrate de usar tu función get_db_connection())
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        # Consulta SQL simplificada usando solo las columnas existentes
        query = """
            SELECT tipo_movimiento, base_legal
            FROM base_legal_accion
            ORDER BY tipo_movimiento ASC
        """

        cursor.execute(query)
        resultados = cursor.fetchall()

        cursor.close()
        conn.close()

        return jsonify(resultados), 200

    except Exception as e:
        print(f"Error en /api/base-legal: {str(e)}")
        return jsonify({"error": "Error al obtener la base legal", "detalles": str(e)}), 500


# =========================
# 💰 ESCALA DE REMUNERACIÓN POR GRUPO OCUPACIONAL
# =========================
@app.route('/api/escala-ocupacional', methods=['GET'])
def get_escala_ocupacional():
    token = request.headers.get('Authorization', '')
    if not decodificar_token(token):
        return jsonify({"error": "No autorizado"}), 403

    try:
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("""
            SELECT grupo_ocupacional, grado, remuneracion
            FROM escala_ocupacional
            WHERE activo = 1
            ORDER BY grupo_ocupacional ASC
        """)
        resultados = cursor.fetchall()

        cursor.close()
        conn.close()

        return jsonify(resultados), 200

    except Exception as e:
        print(f"Error en /api/escala-ocupacional: {str(e)}")
        return jsonify({"error": "Error al obtener la escala ocupacional", "detalles": str(e)}), 500


# =========================
# 🏛️ AUTORIDADES (CRUD)
# =========================
@app.route('/api/autoridades', methods=['GET'])
def obtener_autoridades():
    token = request.headers.get('Authorization', '')
    if not decodificar_token(token):
        return jsonify({"error": "No autorizado"}), 403

    conexion = None
    cursor   = None

    try:
        conexion = get_connection()
        cursor   = conexion.cursor(dictionary=True)

        cursor.execute("""
            SELECT id, nombres, provincia, canton, denominacion_puesto, unidad_organica
            FROM autoridades
            WHERE nombres IS NOT NULL AND nombres != ''
            ORDER BY nombres ASC
        """)
        resultados = cursor.fetchall()
        return jsonify(resultados), 200

    except Exception as e:
        print("ERROR OBTENIENDO AUTORIDADES:", str(e))
        return jsonify({"error": str(e)}), 500

    finally:
        if cursor:  cursor.close()
        if conexion: conexion.close()


@app.route('/api/autoridades', methods=['POST'])
def crear_autoridad():
    if not es_admin():
        return jsonify({"error": "No autorizado"}), 403

    data = request.get_json()
    nombres            = (data.get('nombres') or '').strip().upper()
    denominacion_puesto = (data.get('denominacion_puesto') or '').strip().upper()
    unidad_organica    = (data.get('unidad_organica') or '').strip().upper()
    provincia          = (data.get('provincia') or 'PICHINCHA').strip().upper()
    canton             = (data.get('canton') or 'QUITO').strip().upper()

    if not nombres or not denominacion_puesto:
        return jsonify({"error": "Nombre y puesto son obligatorios"}), 400

    conexion = None
    cursor   = None

    try:
        conexion = get_connection()
        cursor   = conexion.cursor(dictionary=True)

        cursor.execute("""
            INSERT INTO autoridades (nombres, provincia, canton, denominacion_puesto, unidad_organica)
            VALUES (%s, %s, %s, %s, %s)
        """, (nombres, provincia, canton, denominacion_puesto, unidad_organica))

        conexion.commit()
        nuevo_id = cursor.lastrowid
        return jsonify({"message": "Autoridad creada", "id": nuevo_id}), 201

    except Exception as e:
        print("ERROR CREANDO AUTORIDAD:", str(e))
        return jsonify({"error": str(e)}), 500

    finally:
        if cursor:  cursor.close()
        if conexion: conexion.close()


@app.route('/api/autoridades/<int:id>', methods=['PUT'])
def actualizar_autoridad(id):
    if not es_admin():
        return jsonify({"error": "No autorizado"}), 403

    data = request.get_json()
    nombres            = (data.get('nombres') or '').strip().upper()
    denominacion_puesto = (data.get('denominacion_puesto') or '').strip().upper()
    unidad_organica    = (data.get('unidad_organica') or '').strip().upper()
    provincia          = (data.get('provincia') or 'PICHINCHA').strip().upper()
    canton             = (data.get('canton') or 'QUITO').strip().upper()

    if not nombres or not denominacion_puesto:
        return jsonify({"error": "Nombre y puesto son obligatorios"}), 400

    conexion = None
    cursor   = None

    try:
        conexion = get_connection()
        cursor   = conexion.cursor(dictionary=True)

        cursor.execute("""
            UPDATE autoridades
            SET nombres = %s, provincia = %s, canton = %s,
                denominacion_puesto = %s, unidad_organica = %s
            WHERE id = %s
        """, (nombres, provincia, canton, denominacion_puesto, unidad_organica, id))

        conexion.commit()
        return jsonify({"message": "Autoridad actualizada"}), 200

    except Exception as e:
        print("ERROR ACTUALIZANDO AUTORIDAD:", str(e))
        return jsonify({"error": str(e)}), 500

    finally:
        if cursor:  cursor.close()
        if conexion: conexion.close()


@app.route('/api/autoridades/<int:id>', methods=['DELETE'])
def eliminar_autoridad(id):
    if not es_admin():
        return jsonify({"error": "No autorizado"}), 403

    conexion = None
    cursor   = None

    try:
        conexion = get_connection()
        cursor   = conexion.cursor(dictionary=True)

        cursor.execute("DELETE FROM autoridades WHERE id = %s", (id,))
        conexion.commit()
        return jsonify({"message": "Autoridad eliminada"}), 200

    except Exception as e:
        print("ERROR ELIMINANDO AUTORIDAD:", str(e))
        return jsonify({"error": str(e)}), 500

    finally:
        if cursor:  cursor.close()
        if conexion: conexion.close()


# =========================
# 👥 ESTRUCTURA DE PERSONAL (CRUD)
# =========================
@app.route('/api/personal-estructura', methods=['GET'])
def obtener_personal_estructura():
    token = request.headers.get('Authorization', '')
    if not decodificar_token(token):
        return jsonify({"error": "No autorizado"}), 403

    conexion = None
    cursor   = None

    try:
        conexion = get_connection()
        cursor   = conexion.cursor(dictionary=True)

        cursor.execute("""
            SELECT id, nombres, provincia, canton, denominacion_puesto, unidad_organica
            FROM personal_estructura
            WHERE nombres IS NOT NULL AND nombres != ''
            ORDER BY nombres ASC
        """)
        resultados = cursor.fetchall()
        return jsonify(resultados), 200

    except Exception as e:
        print("ERROR OBTENIENDO PERSONAL ESTRUCTURA:", str(e))
        return jsonify({"error": str(e)}), 500

    finally:
        if cursor:  cursor.close()
        if conexion: conexion.close()


@app.route('/api/personal-estructura', methods=['POST'])
def crear_personal_estructura():
    if not es_admin():
        return jsonify({"error": "No autorizado"}), 403

    data = request.get_json()
    nombres             = (data.get('nombres')             or '').strip().upper()
    denominacion_puesto = (data.get('denominacion_puesto') or '').strip().upper()
    unidad_organica     = (data.get('unidad_organica')     or '').strip().upper()
    provincia           = (data.get('provincia')           or 'PICHINCHA').strip().upper()
    canton              = (data.get('canton')              or 'QUITO').strip().upper()

    if not nombres or not denominacion_puesto:
        return jsonify({"error": "Nombre y puesto son obligatorios"}), 400

    conexion = None
    cursor   = None

    try:
        conexion = get_connection()
        cursor   = conexion.cursor(dictionary=True)

        cursor.execute("""
            INSERT INTO personal_estructura (nombres, provincia, canton, denominacion_puesto, unidad_organica)
            VALUES (%s, %s, %s, %s, %s)
        """, (nombres, provincia, canton, denominacion_puesto, unidad_organica))

        conexion.commit()
        return jsonify({"message": "Registro creado", "id": cursor.lastrowid}), 201

    except Exception as e:
        print("ERROR CREANDO PERSONAL ESTRUCTURA:", str(e))
        return jsonify({"error": str(e)}), 500

    finally:
        if cursor:  cursor.close()
        if conexion: conexion.close()


@app.route('/api/personal-estructura/<int:id>', methods=['PUT'])
def actualizar_personal_estructura(id):
    if not es_admin():
        return jsonify({"error": "No autorizado"}), 403

    data = request.get_json()
    nombres             = (data.get('nombres')             or '').strip().upper()
    denominacion_puesto = (data.get('denominacion_puesto') or '').strip().upper()
    unidad_organica     = (data.get('unidad_organica')     or '').strip().upper()
    provincia           = (data.get('provincia')           or 'PICHINCHA').strip().upper()
    canton              = (data.get('canton')              or 'QUITO').strip().upper()

    if not nombres or not denominacion_puesto:
        return jsonify({"error": "Nombre y puesto son obligatorios"}), 400

    conexion = None
    cursor   = None

    try:
        conexion = get_connection()
        cursor   = conexion.cursor(dictionary=True)

        cursor.execute("""
            UPDATE personal_estructura
            SET nombres = %s, provincia = %s, canton = %s,
                denominacion_puesto = %s, unidad_organica = %s
            WHERE id = %s
        """, (nombres, provincia, canton, denominacion_puesto, unidad_organica, id))

        conexion.commit()
        return jsonify({"message": "Registro actualizado"}), 200

    except Exception as e:
        print("ERROR ACTUALIZANDO PERSONAL ESTRUCTURA:", str(e))
        return jsonify({"error": str(e)}), 500

    finally:
        if cursor:  cursor.close()
        if conexion: conexion.close()


@app.route('/api/personal-estructura/<int:id>', methods=['DELETE'])
def eliminar_personal_estructura(id):
    if not es_admin():
        return jsonify({"error": "No autorizado"}), 403

    conexion = None
    cursor   = None

    try:
        conexion = get_connection()
        cursor   = conexion.cursor(dictionary=True)

        cursor.execute("DELETE FROM personal_estructura WHERE id = %s", (id,))
        conexion.commit()
        return jsonify({"message": "Registro eliminado"}), 200

    except Exception as e:
        print("ERROR ELIMINANDO PERSONAL ESTRUCTURA:", str(e))
        return jsonify({"error": str(e)}), 500

    finally:
        if cursor:  cursor.close()
        if conexion: conexion.close()


# =========================
# 📄 GENERAR ACCIÓN DE PERSONAL (EXCEL)
# =========================
@app.route('/api/generar-accion', methods=['POST'])
def generar_accion():
    if not es_admin():
        return jsonify({"error": "No autorizado"}), 403

    if not OPENPYXL_DISPONIBLE:
        return jsonify({"error": "Librería openpyxl no instalada. Ejecute: pip install openpyxl"}), 503

    datos = request.get_json()
    if not datos:
        return jsonify({"error": "Datos requeridos"}), 400

    print(">>> generar-accion recibido:", {
        "apellidos":        datos.get("apellidos"),
        "cedula":           datos.get("cedula"),
        "fecha_rige_desde": datos.get("fecha_rige_desde"),
        "fecha_rige_hasta": datos.get("fecha_rige_hasta"),
        "tipo_accion":      datos.get("tipo_accion"),
    })

    carpeta = os.path.join(os.path.dirname(__file__), 'plantillas')
    archivos = glob.glob(os.path.join(carpeta, '*.xlsx'))

    if not archivos:
        return jsonify({"error": "No se encontró ningún archivo .xlsx en la carpeta 'plantillas/'. Coloque su plantilla allí."}), 404

    ruta_plantilla = archivos[0]

    try:
        from openpyxl.utils import coordinate_to_tuple
        from datetime import datetime as dt

        def escribir_celda(ws, addr, valor):
            """Escribe en la celda ancla del rango fusionado que contiene addr."""
            row, col = coordinate_to_tuple(addr)
            for rango in ws.merged_cells.ranges:
                if rango.min_row <= row <= rango.max_row and rango.min_col <= col <= rango.max_col:
                    ws.cell(row=rango.min_row, column=rango.min_col, value=valor)
                    return
            ws.cell(row=row, column=col, value=valor)

        def parse_fecha(s, como_texto=False):
            """Convierte string YYYY-MM-DD a datetime o texto DD/MM/YYYY."""
            if not s:
                return ''
            for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y'):
                try:
                    d = dt.strptime(s, fmt)
                    return d.strftime('%d/%m/%Y') if como_texto else d
                except ValueError:
                    pass
            return s

        wb = openpyxl.load_workbook(ruta_plantilla)
        hoja = wb['ap'] if 'ap' in wb.sheetnames else wb.active

        # ── ENCABEZADO ──────────────────────────────────────────────────────
        # M3 = número de AP  |  K5 = fecha de elaboración (ya tiene formato de fecha en plantilla)
        escribir_celda(hoja, 'M3', datos.get('numero_accion', ''))
        escribir_celda(hoja, 'K5', parse_fecha(datos.get('fecha_elaboracion', '')))

        # ── APELLIDOS / NOMBRES (A6:H7 y I6:P7) ─────────────────────────────
        # A8 e I8 son los LABELS; los datos van en A6 e I6
        escribir_celda(hoja, 'A6', datos.get('apellidos', ''))
        escribir_celda(hoja, 'I6', datos.get('nombres', ''))

        # ── CÉDULA Y FECHAS DE VIGENCIA ──────────────────────────────────────
        # A11:D11 = label CÉDULA  →  E11:H11 = valor cédula
        # I11:L11 = fecha rige desde  |  M11:P11 = fecha rige hasta
        # Las fechas se escriben como texto DD/MM/YYYY para evitar números seriales
        escribir_celda(hoja, 'E11', datos.get('cedula', ''))
        escribir_celda(hoja, 'I11', parse_fecha(datos.get('fecha_rige_desde', ''), como_texto=True))
        escribir_celda(hoja, 'M11', parse_fecha(datos.get('fecha_rige_hasta', ''), como_texto=True))

        # ── TIPO DE ACCIÓN (casillas X) ─────────────────────────────────────────────────────
        import unicodedata as _ud

        def _norm(t):
            """Normaliza texto: quita tildes, colapsa espacios, pone en mayúsculas."""
            t = str(t or '').strip().upper()
            t = _ud.normalize('NFD', t)
            t = ''.join(c for c in t if _ud.category(c) != 'Mn')
            return ' '.join(t.split())

        # Mapa de acciones normalizadas → coordenada Excel
        MARCAS_ACCION = {
            'INGRESO':               'A14',
            'REINGRESO':             'A15',
            'RESTITUCION':           'A16',
            'REINTEGRO':             'A17',
            'ASCENSO':               'A18',
            'TRASLADO':              'A19',
            'TRASPASO':              'D14',
            'CAMBIO ADMINISTRATIVO': 'D15',
            'INTERCAMBIO VOLUNTARIO':'D16',
            'LICENCIA':              'D17',
            'COMISION DE SERVICIOS': 'D18',
            'SANCIONES':             'D19',
            'INCREMENTO RMU':        'I14',
            'SUBROGACION':           'I15',
            'ENCARGO':               'I16',
            'CESACION DE FUNCIONES': 'I17',
            'DESTITUCION':           'I18',
            'VACACIONES':            'I19',
            'REVISION CLAS. PUESTO': 'L14',
            'OTRO':                  'L15',
        }

        tipo_raw  = datos.get('tipo_accion') or datos.get('accion_personal') or ''
        tipo_norm = _norm(tipo_raw)

        celda_marca = MARCAS_ACCION.get(tipo_norm)

        if celda_marca:
            from openpyxl.utils import coordinate_to_tuple as _ctt
            from openpyxl.styles import Font as _Font
            escribir_celda(hoja, celda_marca, 'X')
            _r, _c = _ctt(celda_marca)
            _celda = hoja.cell(row=_r, column=_c)
            _celda.font = _Font(bold=True, name=_celda.font.name, size=_celda.font.size)
        else:
            # Fallback: si no coincide con ninguna opción, marcar OTRO
            from openpyxl.utils import coordinate_to_tuple as _ctt
            from openpyxl.styles import Font as _Font
            escribir_celda(hoja, 'L15', 'X')
            _r, _c = _ctt('L15')
            _celda = hoja.cell(row=_r, column=_c)
            _celda.font = _Font(bold=True, name=_celda.font.name, size=_celda.font.size)
            if tipo_norm not in ('', 'OTRO'):
                escribir_celda(hoja, 'L16', tipo_raw.upper())

        # ── MOTIVACIÓN / BASE LEGAL ───────────────────────────────────────────
        # A24:O24 = celda grande de motivación
        escribir_celda(hoja, 'A24', datos.get('motivo_legal', ''))

        # ── SITUACIÓN ACTUAL ──────────────────────────────────────────────────
        proc_actual  = datos.get('proceso_institucional_actual', '')
        unidad_act   = datos.get('unidad_administrativa', '')
        lugar_act    = datos.get('lugar_trabajo_actual', '') or datos.get('ciudad', '')
        denom_act    = datos.get('denominacion_actual', '')  or datos.get('cargo', '')
        grupo        = datos.get('grupo_ocupacional', '')
        partida_act  = datos.get('partida_actual', '')

        nivel_gest_act  = datos.get('nivel_gestion_actual', '')
        nivel_gest_prop = datos.get('nivel_gestion_propuesta', '') or nivel_gest_act

        escribir_celda(hoja, 'B28', proc_actual)
        escribir_celda(hoja, 'B30', nivel_gest_act)
        escribir_celda(hoja, 'B32', unidad_act)
        escribir_celda(hoja, 'B34', lugar_act)
        escribir_celda(hoja, 'B36', denom_act)
        escribir_celda(hoja, 'B38', grupo)
        # B40 y B42 son fórmulas VLOOKUP sobre B38 — se recalculan solas al abrir Excel
        escribir_celda(hoja, 'B44', partida_act)

        # ── SITUACIÓN PROPUESTA (hereda actual si no se especifica) ───────────
        proc_prop   = datos.get('proceso_institucional_propuesta', '') or proc_actual
        unidad_prop = datos.get('unidad_administrativa_propuesta', '') or unidad_act
        lugar_prop  = datos.get('lugar_trabajo_propuesta', '')  or lugar_act
        denom_prop  = datos.get('denominacion_propuesta', '')   or denom_act
        partida_prop = datos.get('partida_propuesta', '')       or partida_act

        escribir_celda(hoja, 'J28', proc_prop)
        escribir_celda(hoja, 'J30', nivel_gest_prop)
        escribir_celda(hoja, 'J32', unidad_prop)
        escribir_celda(hoja, 'J34', lugar_prop)
        escribir_celda(hoja, 'J36', denom_prop)
        escribir_celda(hoja, 'J38', grupo)
        # J40 y J42 son fórmulas VLOOKUP sobre J38 — se recalculan solas al abrir Excel
        escribir_celda(hoja, 'J44', partida_prop)

        # ── POSESIÓN DEL PUESTO ───────────────────────────────────────────────
        # C48 = "YO, [nombre del servidor]" en la sección de posesión del cargo
        nombre_posesion = f"{datos.get('apellidos', '')} {datos.get('nombres', '')}".strip()
        if nombre_posesion:
            escribir_celda(hoja, 'C48', nombre_posesion)
        # N48 = número de documento de identificación (cédula)
        if datos.get('cedula'):
            escribir_celda(hoja, 'N48', datos.get('cedula'))
        escribir_celda(hoja, 'C50', datos.get('ciudad', ''))

        # ── RESPONSABLES DE APROBACIÓN ────────────────────────────────────────
        # Director TH: nombre en C61:G61, puesto en C62:G62
        escribir_celda(hoja, 'C61', datos.get('nombre_director_th', ''))
        escribir_celda(hoja, 'C62', datos.get('puesto_director_th', ''))
        # Autoridad nominadora: nombre en K61:O61, puesto en K62:O62
        escribir_celda(hoja, 'K61', datos.get('nombre_autoridad', ''))
        escribir_celda(hoja, 'K62', datos.get('puesto_autoridad', ''))

        # Responsable de Talento Humano: nombre en C61:E61, puesto en C62:E62
        escribir_celda(hoja, 'C61', datos.get('nombre_responsable_th', ''))
        escribir_celda(hoja, 'C62', datos.get('puesto_responsable_th', ''))

        # ── ACEPTACIÓN DEL SERVIDOR ───────────────────────────────────────────
        # C74 normalmente tiene fórmula =A6&" "&I6 — sobreescribimos con nombre completo
        escribir_celda(hoja, 'C74', datos.get('aceptacion_servidor', ''))
        # C75 normalmente tiene fórmula =+K5 — sobreescribimos con la fecha de aceptación
        fecha_acep = parse_fecha(
            datos.get('fecha_aceptacion', '') or datos.get('fecha_elaboracion', ''),
            como_texto=True
        )
        if fecha_acep:
            escribir_celda(hoja, 'C75', fecha_acep)

        # ── RESPONSABLES ELABORACIÓN / REVISIÓN / REGISTRO ───────────────────
        # Elaboración: C87:E87 (nombre) y C88:E88 (puesto)
        escribir_celda(hoja, 'C87', datos.get('elaborado_por', ''))
        escribir_celda(hoja, 'C88', datos.get('puesto_elaborado', ''))
        # Revisión: G87:K87 (nombre) y G88:K88 (puesto)
        escribir_celda(hoja, 'G87', datos.get('revisado_por', ''))
        escribir_celda(hoja, 'G88', datos.get('puesto_revisado', ''))
        # Registro: M87:O87 (nombre) y M88:O88 (puesto)
        escribir_celda(hoja, 'M87', datos.get('registrado_por', ''))
        escribir_celda(hoja, 'M88', datos.get('puesto_registrado', ''))

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        apellidos = datos.get('apellidos', 'Generada').replace(' ', '_')
        nombre_archivo = f"Accion_Personal_{apellidos}.xlsx"

        registrar_auditoria(
            usuario=obtener_usuario(),
            accion='EXPORT',
            tabla='accion_personal',
            registro_id=None,
            antes=None,
            despues={
                "cedula":    datos.get('cedula', ''),
                "apellidos": datos.get('apellidos', ''),
                "tipo":      datos.get('tipo_accion', ''),
                "desde":     datos.get('fecha_rige_desde', ''),
                "hasta":     datos.get('fecha_rige_hasta', ''),
            },
            detalle=f"Generación de Acción de Personal — {datos.get('apellidos', '')} {datos.get('cedula', '')}"
        )

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nombre_archivo
        )

    except Exception as e:
        print("ERROR GENERANDO ACCION EXCEL:", str(e))
        return jsonify({"error": f"Error al procesar la plantilla: {str(e)}"}), 500


# =========================
# 📁 HISTORIAL DE ACCIONES DE PERSONAL
# =========================
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'uploads', 'acciones')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

EXTENSIONES_PERMITIDAS = {'pdf', 'xlsx', 'xls', 'docx', 'doc'}

def extension_permitida(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in EXTENSIONES_PERMITIDAS

def nombre_seguro(filename):
    import re
    filename = os.path.basename(filename)
    filename = re.sub(r'[^A-Za-z0-9._\-]', '_', filename)
    return filename or 'archivo'


@app.route('/api/historial-acciones/buscar', methods=['GET'])
def buscar_historial_acciones():
    token = request.headers.get('Authorization', '')
    if not decodificar_token(token):
        return jsonify({"error": "No autorizado"}), 403

    q = (request.args.get('q') or '').strip()
    if not q:
        return jsonify({"error": "Parámetro de búsqueda requerido"}), 400

    conexion = None
    cursor   = None

    try:
        conexion = get_connection()
        cursor   = conexion.cursor(dictionary=True)

        # Buscar persona en tabla personal (por cédula, nombre o nro)
        cursor.execute("""
            SELECT id, nro AS numero_nomina, cedula, nombres, cargo, unidad, modalidad, rmu
            FROM personal
            WHERE cedula = %s
               OR UPPER(nombres) LIKE UPPER(%s)
               OR nro = %s
            LIMIT 1
        """, (q, f'%{q}%', q))
        persona = cursor.fetchone()

        # Determinar cédula para buscar acciones
        cedula_busqueda = None
        if persona:
            cedula_busqueda = persona['cedula']
        else:
            # Si no está en personal activo, buscar en historial directamente
            cursor.execute("""
                SELECT cedula, nombres FROM historial_acciones
                WHERE cedula = %s OR UPPER(nombres) LIKE UPPER(%s)
                LIMIT 1
            """, (q, f'%{q}%'))
            hist_ref = cursor.fetchone()
            if hist_ref:
                cedula_busqueda = hist_ref['cedula']
                persona = { 'cedula': hist_ref['cedula'], 'nombres': hist_ref['nombres'],
                            'cargo': None, 'unidad': None, 'encontrado_en_personal': False }

        if not cedula_busqueda and not persona:
            return jsonify({"persona": None, "acciones": []}), 200

        if persona and 'encontrado_en_personal' not in persona:
            persona['encontrado_en_personal'] = True

        # Buscar acciones del sistema antiguo
        cursor.execute("""
            SELECT id, cedula, nombres, numero_accion, tipo_accion,
                   fecha_accion, fecha_registro, archivo_nombre, registrado_por
            FROM historial_acciones
            WHERE cedula = %s
        """, (cedula_busqueda,))
        acciones_viejas = cursor.fetchall()

        for a in acciones_viejas:
            a['es_nativo'] = False
            a['estado_documento'] = None

        # Buscar acciones del nuevo sistema (nativas)
        cursor.execute("""
            SELECT id, cedula_funcionario as cedula, numero_accion, estado as estado_documento,
                   fecha_creacion as fecha_registro, fecha_creacion as fecha_accion,
                   usuario_creacion as registrado_por, ruta_pdf_actual,
                   datos_formulario
            FROM documentos_accion_personal
            WHERE cedula_funcionario = %s AND estado != 'BORRADOR'
        """, (cedula_busqueda,))
        acciones_nuevas = cursor.fetchall()

        import json
        import os

        for a in acciones_nuevas:
            a['es_nativo'] = True
            a['nombres'] = persona['nombres'] if persona else ''

            # Extraer tipo_accion del JSON si es posible
            try:
                datos = json.loads(a.get('datos_formulario') or '{}')
                a['tipo_accion'] = datos.get('tipo_accion') or datos.get('accion_personal') or 'ACCIÓN DE PERSONAL'
                a['datos_formulario'] = datos
            except:
                a['tipo_accion'] = 'ACCIÓN DE PERSONAL'
                a['datos_formulario'] = {}

            a['archivo_nombre'] = os.path.basename(a['ruta_pdf_actual']) if a.get('ruta_pdf_actual') else None

            # Limpiar datos crudos
            if 'ruta_pdf_actual' in a: del a['ruta_pdf_actual']

        acciones = acciones_viejas + acciones_nuevas

        # Ordenar por fecha_registro descendente
        from datetime import datetime, date
        acciones.sort(
            key=lambda x: x.get('fecha_registro') if isinstance(x.get('fecha_registro'), datetime) else datetime.min,
            reverse=True
        )

        # Serializar fechas
        for a in acciones:
            if isinstance(a.get('fecha_accion'), date):
                a['fecha_accion'] = a['fecha_accion'].isoformat()
            if isinstance(a.get('fecha_registro'), datetime):
                a['fecha_registro'] = a['fecha_registro'].strftime('%Y-%m-%d %H:%M:%S')

        return jsonify({"persona": persona, "acciones": acciones}), 200

    except Exception as e:
        print("ERROR BUSCANDO HISTORIAL:", str(e))
        return jsonify({"error": str(e)}), 500

    finally:
        if cursor:  cursor.close()
        if conexion: conexion.close()


@app.route('/api/historial-acciones/subir', methods=['POST'])
def subir_accion():
    if not es_admin():
        return jsonify({"error": "No autorizado"}), 403

    cedula      = (request.form.get('cedula')       or '').strip()
    nombres     = (request.form.get('nombres')      or '').strip().upper()
    num_accion  = (request.form.get('numero_accion') or '').strip()
    tipo_accion = (request.form.get('tipo_accion')  or '').strip()
    fecha_accion = request.form.get('fecha_accion') or None
    registrado_por = obtener_usuario()

    if not cedula or not nombres:
        return jsonify({"error": "Cédula y nombres son obligatorios"}), 400

    archivo = request.files.get('archivo')
    archivo_nombre = None
    archivo_path   = None

    if archivo and archivo.filename:
        if not extension_permitida(archivo.filename):
            return jsonify({"error": "Extensión no permitida. Use PDF, Excel o Word"}), 400

        import time
        ts = int(time.time())
        ext = archivo.filename.rsplit('.', 1)[1].lower()
        archivo_nombre = f"{cedula}_{ts}.{ext}"
        archivo_path   = os.path.join(UPLOAD_FOLDER, nombre_seguro(archivo_nombre))
        archivo.save(archivo_path)

    conexion = None
    cursor   = None

    try:
        conexion = get_connection()
        cursor   = conexion.cursor(dictionary=True)

        cursor.execute("""
            INSERT INTO historial_acciones
                (cedula, nombres, numero_accion, tipo_accion, fecha_accion, archivo_nombre, archivo_path, registrado_por)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (cedula, nombres, num_accion, tipo_accion, fecha_accion or None,
              archivo_nombre, archivo_path, registrado_por))

        conexion.commit()
        return jsonify({"message": "Acción registrada", "id": cursor.lastrowid}), 201

    except Exception as e:
        # Borrar archivo si falla el INSERT
        if archivo_path and os.path.exists(archivo_path):
            os.remove(archivo_path)
        print("ERROR SUBIENDO ACCION:", str(e))
        return jsonify({"error": str(e)}), 500

    finally:
        if cursor:  cursor.close()
        if conexion: conexion.close()


@app.route('/api/historial-acciones/<int:id>/descargar', methods=['GET'])
def descargar_accion(id):
    token = request.headers.get('Authorization', '')
    if not decodificar_token(token):
        return jsonify({"error": "No autorizado"}), 403

    conexion = None
    cursor   = None

    try:
        conexion = get_connection()
        cursor   = conexion.cursor(dictionary=True)

        cursor.execute("SELECT archivo_path, archivo_nombre FROM historial_acciones WHERE id = %s", (id,))
        row = cursor.fetchone()

        if not row or not row['archivo_path']:
            return jsonify({"error": "Archivo no encontrado"}), 404

        if not os.path.exists(row['archivo_path']):
            return jsonify({"error": "El archivo ya no existe en el servidor"}), 404

        ext = row['archivo_nombre'].rsplit('.', 1)[-1].lower() if row['archivo_nombre'] else 'bin'
        mime_map = {
            'pdf':  'application/pdf',
            'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'xls':  'application/vnd.ms-excel',
            'docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            'doc':  'application/msword'
        }
        mimetype = mime_map.get(ext, 'application/octet-stream')

        return send_file(
            row['archivo_path'],
            as_attachment=True,
            download_name=row['archivo_nombre'] or f'accion_{id}.{ext}',
            mimetype=mimetype
        )

    except Exception as e:
        print("ERROR DESCARGANDO ACCION:", str(e))
        return jsonify({"error": str(e)}), 500

    finally:
        if cursor:  cursor.close()
        if conexion: conexion.close()


@app.route('/api/historial-acciones/<int:id>', methods=['DELETE'])
def eliminar_accion_historial(id):
    if not es_admin():
        return jsonify({"error": "No autorizado"}), 403

    conexion = None
    cursor   = None

    try:
        conexion = get_connection()
        cursor   = conexion.cursor(dictionary=True)

        cursor.execute("SELECT archivo_path FROM historial_acciones WHERE id = %s", (id,))
        row = cursor.fetchone()

        if row and row['archivo_path'] and os.path.exists(row['archivo_path']):
            os.remove(row['archivo_path'])

        cursor.execute("DELETE FROM historial_acciones WHERE id = %s", (id,))
        conexion.commit()
        return jsonify({"message": "Acción eliminada"}), 200

    except Exception as e:
        print("ERROR ELIMINANDO ACCION:", str(e))
        return jsonify({"error": str(e)}), 500

    finally:
        if cursor:  cursor.close()
        if conexion: conexion.close()




# =========================
# 🚀 RUTAS DE FIRMAS ELECTRONICAS (NUEVAS)
# =========================

@app.route('/api/acciones-personal', methods=['POST'])
def guardar_borrador():
    token = request.headers.get('Authorization', '')
    if not decodificar_token(token):
        return jsonify({"error": "No autorizado"}), 403

    datos = request.json
    numero_accion = datos.get('numero_accion')
    cedula = datos.get('cedula')

    if not numero_accion or not cedula:
        return jsonify({"ok": False, "error": "Número de acción y cédula son obligatorios"}), 400

    usuario = obtener_usuario()
    exito, doc_id, error = DocumentoService.guardar_borrador(
        numero_accion, cedula, datos, usuario
    )

    if exito:
        return jsonify({
            "ok": True,
            "mensaje": "Borrador guardado",
            "doc_id": doc_id,
            "numero_accion": numero_accion,
            "estado": "BORRADOR"
        })
    return jsonify({"ok": False, "error": error, "doc_id": doc_id}), 400


@app.route('/api/acciones-personal/<int:doc_id>/preparar-firmas', methods=['POST'])
def preparar_firmas(doc_id):
    """
    Prepara un documento para firmas:
    1. Obtiene el documento y valida que esté en BORRADOR
    2. Genera el Excel con los datos del formulario guardado
    3. Convierte el Excel a PDF
    4. Crea las firmas pendientes
    5. Bloquea el documento
    """
    token = request.headers.get('Authorization', '')
    if not decodificar_token(token):
        return jsonify({"ok": False, "error": "No autorizado"}), 403

    try:
        # 1. Obtener documento
        doc = DocumentoService.obtener_documento(doc_id)
        if not doc:
            return jsonify({"ok": False, "error": "Documento no encontrado"}), 404

        if doc['estado'] != 'BORRADOR':
            return jsonify({"ok": False, "error": f"El documento ya no está en borrador (estado: {doc['estado']})"}), 400

        # 2. Recuperar datos del formulario
        datos_formulario = json.loads(doc['datos_formulario']) if doc['datos_formulario'] else {}
        if not datos_formulario:
            return jsonify({"ok": False, "error": "El documento no tiene datos de formulario guardados"}), 400

        # 3. Crear directorio del documento
        doc_dir = os.path.join(config_firmas.BASE_STORAGE_DIR, str(doc_id))
        excel_dir = os.path.join(doc_dir, 'excel')
        pdf_dir = os.path.join(doc_dir, 'pdf')
        os.makedirs(excel_dir, exist_ok=True)
        os.makedirs(pdf_dir, exist_ok=True)

        # 4. Generar Excel con datos reales (reutilizando lógica de generar_accion)
        excel_path = _generar_excel_documento(datos_formulario, excel_dir, doc['numero_accion'])

        # 5. Convertir a PDF
        pdf_path = PdfConverter.convert_excel_to_pdf(
            excel_path, pdf_dir, config_firmas.LIBREOFFICE_PATH
        )

        if not pdf_path or not os.path.exists(pdf_path):
            return jsonify({
                "ok": False,
                "codigo": "PDF_NO_GENERADO",
                "error": "LibreOffice no generó el PDF de la Acción de Personal."
            }), 500

        total_paginas = obtener_total_paginas_pdf(pdf_path)
        app.logger.info(
            "Documento %s preparado con %s página(s): %s",
            doc_id, total_paginas, pdf_path
        )

        # 6. Calcular hash inicial
        hash_inicial = FirmaService.calcular_hash(pdf_path)

        # 7. Actualizar documento en BD
        DocumentoService.actualizar_estado_documento(
            doc_id,
            estado='PENDIENTE_FIRMAS',
            ruta_excel=excel_path,
            ruta_pdf_original=pdf_path,
            ruta_pdf_actual=pdf_path,
            hash_pdf=hash_inicial,
            bloqueado=True
        )

        # 8. Registrar versión original del PDF
        conn = DocumentoService.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO versiones_documento
            (documento_id, numero_version, tipo_version, ruta_archivo, hash_archivo)
            VALUES (%s, 1, 'ORIGINAL', %s, %s)
        """, (doc_id, pdf_path, hash_inicial))
        cursor.execute(
            "UPDATE documentos_accion_personal SET version_documento = 1 WHERE id = %s",
            (doc_id,)
        )
        conn.commit()
        cursor.close()
        conn.close()

        # 9. Crear firmas pendientes
        exito, error = DocumentoService.crear_firmas_pendientes(
            doc_id, config_firmas.SECCIONES_FIRMA, datos_formulario
        )
        if not exito:
            return jsonify({"ok": False, "error": f"Error creando firmas: {error}"}), 500

        # 10. Obtener firmas creadas para devolver
        firmas = DocumentoService.obtener_firmas_documento(doc_id)

        return jsonify({
            "ok": True,
            "mensaje": "Documento preparado para firmas",
            "documento_id": doc_id,
            "numero_accion": doc['numero_accion'],
            "estado": "PENDIENTE_FIRMAS",
            "excel_disponible": True,
            "pdf_disponible": True,
            "firmas": firmas
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route('/api/acciones-personal/<int:doc_id>/firmas', methods=['GET'])
def listar_firmas(doc_id):
    """Lista todas las firmas de un documento, con su estado."""
    token = request.headers.get('Authorization', '')
    if not decodificar_token(token):
        return jsonify({"error": "No autorizado"}), 403

    firmas = DocumentoService.obtener_firmas_documento(doc_id)
    return jsonify(firmas)


@app.route('/api/acciones-personal/<int:doc_id>/firmar', methods=['POST'])
def firmar_documento(doc_id):
    """
    Firma una sección del documento con certificado PKCS#12 (.p12/.pfx).

    Las firmas son independientes: no se valida orden secuencial. La página
    configurada se ajusta automáticamente al número real de páginas del PDF.
    """
    token = request.headers.get('Authorization', '')
    if not decodificar_token(token):
        return jsonify({"ok": False, "error": "No autorizado"}), 403

    p12_bytes = None
    password = None
    nuevo_pdf_path = None

    try:
        # 1. Validaciones de entrada
        cert_file = request.files.get('certificado')
        password = request.form.get('password', '')
        seccion = (request.form.get('seccion') or '').strip()

        if not cert_file or not cert_file.filename:
            return jsonify({
                'ok': False,
                'codigo': 'CERTIFICADO_NO_VALIDO',
                'error': 'Seleccione un certificado .p12 o .pfx.'
            }), 400

        if not password:
            return jsonify({
                'ok': False,
                'codigo': 'CERTIFICADO_PASSWORD_INVALIDO',
                'error': 'Ingrese la contraseña del certificado.'
            }), 400

        if not seccion:
            return jsonify({
                'ok': False,
                'codigo': 'SECCION_REQUERIDA',
                'error': 'Seleccione la sección a firmar.'
            }), 400

        filename = os.path.basename(cert_file.filename)
        if not filename.lower().endswith(('.p12', '.pfx')):
            return jsonify({
                'ok': False,
                'codigo': 'CERTIFICADO_NO_VALIDO',
                'error': 'El archivo debe tener extensión .p12 o .pfx.'
            }), 400

        p12_bytes = cert_file.read(MAX_CERTIFICADO_BYTES + 1)

        if not p12_bytes:
            return jsonify({
                'ok': False,
                'codigo': 'CERTIFICADO_VACIO',
                'error': 'El archivo del certificado está vacío.'
            }), 400

        if len(p12_bytes) > MAX_CERTIFICADO_BYTES:
            return jsonify({
                'ok': False,
                'codigo': 'CERTIFICADO_DEMASIADO_GRANDE',
                'error': 'El certificado no debe superar el tamaño permitido.'
            }), 413

        if seccion not in config_firmas.SECCIONES_FIRMA:
            return jsonify({
                'ok': False,
                'codigo': 'SECCION_NO_VALIDA',
                'error': f'Sección no válida: {seccion}'
            }), 400

        # 2. Cargar y validar certificado
        cert_info = CertificadoService.cargar_certificado(p12_bytes, password)
        if not cert_info.get('valido'):
            return jsonify({
                'ok': False,
                'codigo': 'CERTIFICADO_PASSWORD_INVALIDO',
                'error': cert_info.get(
                    'error',
                    'Certificado inválido o contraseña incorrecta.'
                )
            }), 400

        vigente, mensaje_vigencia = CertificadoService.validar_vigencia(cert_info)
        if not vigente:
            return jsonify({
                'ok': False,
                'codigo': 'CERTIFICADO_VENCIDO',
                'error': mensaje_vigencia
            }), 422

        # 3. Obtener y validar documento
        doc = DocumentoService.obtener_documento(doc_id)
        if not doc:
            return jsonify({
                'ok': False,
                'codigo': 'DOCUMENTO_NO_ENCONTRADO',
                'error': 'Documento no encontrado.'
            }), 404

        if doc.get('estado') not in ('PENDIENTE_FIRMAS', 'FIRMADO_PARCIALMENTE'):
            return jsonify({
                'ok': False,
                'codigo': 'DOCUMENTO_BLOQUEADO',
                'error': (
                    'El documento no acepta firmas '
                    f"(estado: {doc.get('estado')})."
                )
            }), 409

        firmas = DocumentoService.obtener_firmas_documento(doc_id)
        firma_seccion = next(
            (firma for firma in firmas if firma.get('seccion') == seccion),
            None
        )

        if not firma_seccion:
            return jsonify({
                'ok': False,
                'codigo': 'FIRMA_NO_ENCONTRADA',
                'error': 'La sección no está configurada para este documento.'
            }), 404

        if str(firma_seccion.get('estado', '')).upper() == 'FIRMADA':
            return jsonify({
                'ok': False,
                'codigo': 'FIRMA_YA_REALIZADA',
                'error': 'Esta sección ya fue firmada.'
            }), 409

        # No se comprueban firmas anteriores: todas son independientes.

        # 4. Validar el PDF actual y corregir página de firma
        pdf_base = doc.get('ruta_pdf_actual') or doc.get('ruta_pdf_original')
        if not pdf_base or not os.path.exists(pdf_base):
            return jsonify({
                'ok': False,
                'codigo': 'PDF_NO_ENCONTRADO',
                'error': 'El PDF actual del documento no existe.'
            }), 404

        total_paginas = obtener_total_paginas_pdf(pdf_base)
        conf_seccion = config_firmas.SECCIONES_FIRMA[seccion]

        posicion_original = (
            conf_seccion.get('posicion')
            if isinstance(conf_seccion, dict)
            else conf_seccion
        )

        pagina_configurada = 0
        if isinstance(posicion_original, dict):
            pagina_configurada = posicion_original.get('pagina',
                posicion_original.get('page',
                    posicion_original.get('on_page', 0)
                )
            )

        if posicion_original is None:
            return jsonify({
                'ok': False,
                'codigo': 'POSICION_FIRMA_NO_CONFIGURADA',
                'error': f'No existe posición de firma para {seccion}.'
            }), 422

        posicion_firma, pagina_firma = normalizar_posicion_firma(
            posicion_original,
            total_paginas,
            pagina_configurada
        )

        app.logger.info(
            'Firma documento=%s seccion=%s paginas=%s pagina_configurada=%s pagina_usada=%s',
            doc_id,
            seccion,
            total_paginas,
            pagina_configurada,
            pagina_firma
        )

        # 5. Preparar versión de salida
        hash_antes = FirmaService.calcular_hash(pdf_base)
        version = int(doc.get('version_documento') or 1) + 1
        pdf_dir = os.path.dirname(pdf_base)
        os.makedirs(pdf_dir, exist_ok=True)

        numero_accion_safe = str(
            doc.get('numero_accion') or 'documento'
        ).replace(' ', '_').replace('/', '-')
        nuevo_pdf_path = os.path.join(
            pdf_dir,
            f'{numero_accion_safe}-v{version}.pdf'
        )

        # 6. Aplicar firma
        try:
            exito_firma, error_firma = ejecutar_firma_pdf(
                pdf_base,
                nuevo_pdf_path,
                cert_info,
                seccion,
                firma_seccion.get('nombre_firmante', ''),
                firma_seccion.get('cargo_firmante', ''),
                posicion_firma,
                pagina_firma
            )
        except Exception as error:
            detalle = str(error)
            app.logger.exception(
                'Error aplicando firma PDF documento=%s seccion=%s',
                doc_id,
                seccion
            )

            if 'Page index out of range' in detalle:
                return jsonify({
                    'ok': False,
                    'codigo': 'PAGINA_FIRMA_NO_VALIDA',
                    'error': (
                        'La página configurada para la firma no existe en el PDF. '
                        f'El documento tiene {total_paginas} página(s) y se intentó '
                        f'usar el índice {pagina_firma}. Revise services/firma_service.py.'
                    )
                }), 422

            return jsonify({
                'ok': False,
                'codigo': 'ERROR_FIRMA_PDF',
                'error': f'No se pudo aplicar la firma: {detalle}'
            }), 500

        if not exito_firma:
            detalle = str(error_firma or 'Error desconocido al firmar el PDF.')

            if nuevo_pdf_path and os.path.exists(nuevo_pdf_path):
                try:
                    os.remove(nuevo_pdf_path)
                except OSError:
                    pass

            codigo = (
                'PAGINA_FIRMA_NO_VALIDA'
                if 'Page index out of range' in detalle
                else 'ERROR_FIRMA_PDF'
            )
            estado_http = 422 if codigo == 'PAGINA_FIRMA_NO_VALIDA' else 500

            return jsonify({
                'ok': False,
                'codigo': codigo,
                'error': f'Error al firmar: {detalle}'
            }), estado_http

        if not os.path.exists(nuevo_pdf_path) or os.path.getsize(nuevo_pdf_path) == 0:
            return jsonify({
                'ok': False,
                'codigo': 'PDF_FIRMADO_NO_GENERADO',
                'error': 'El servicio no generó el PDF firmado.'
            }), 500

        # 7. Registrar firma en base de datos
        hash_despues = FirmaService.calcular_hash(nuevo_pdf_path)
        exito_bd, firma_id, error_bd = DocumentoService.registrar_firma(
            doc_id,
            seccion,
            cert_info,
            hash_antes,
            hash_despues,
            nuevo_pdf_path,
            version
        )

        if not exito_bd:
            try:
                os.remove(nuevo_pdf_path)
            except OSError:
                pass

            return jsonify({
                'ok': False,
                'codigo': 'ERROR_BASE_DATOS',
                'error': f'No se pudo registrar la firma: {error_bd}'
            }), 500

        doc_actualizado = DocumentoService.obtener_documento(doc_id) or {}

        return jsonify({
            'ok': True,
            'mensaje': 'Documento firmado exitosamente.',
            'documento_id': doc_id,
            'firma_id': firma_id,
            'seccion': seccion,
            'estado_firma': 'FIRMADA',
            'estado_documento': doc_actualizado.get(
                'estado', 'FIRMADO_PARCIALMENTE'
            ),
            'version': version,
            'pagina_firma': pagina_firma,
            'total_paginas': total_paginas,
            'certificado': {
                'titular': cert_info.get('nombre_titular', ''),
                'emisor': cert_info.get('emisor', '')
            }
        }), 200

    except ValueError as error:
        app.logger.exception(
            'PDF inválido al firmar documento=%s', doc_id
        )
        return jsonify({
            'ok': False,
            'codigo': 'PDF_NO_VALIDO',
            'error': str(error)
        }), 422

    except Exception as error:
        app.logger.exception(
            'Error inesperado firmando documento=%s', doc_id
        )
        return jsonify({
            'ok': False,
            'codigo': 'ERROR_INTERNO_FIRMA',
            'error': f'No se pudo firmar el documento: {error}'
        }), 500

    finally:
        # No guardar ni reutilizar datos sensibles del PKCS#12.
        password = None
        p12_bytes = None


@app.route('/api/acciones-personal/<int:doc_id>/finalizar', methods=['POST'])
def finalizar_documento(doc_id):
    """Finaliza un documento si todas las firmas obligatorias están completadas."""
    token = request.headers.get('Authorization', '')
    if not decodificar_token(token):
        return jsonify({"ok": False, "error": "No autorizado"}), 403

    exito, error = DocumentoService.verificar_y_finalizar(doc_id)
    if exito:
        return jsonify({
            'ok': True,
            'mensaje': 'Documento finalizado exitosamente',
            'estado': 'FIRMADO_COMPLETAMENTE'
        })
    return jsonify({'ok': False, 'error': error}), 400


@app.route('/api/acciones-personal/<int:doc_id>/pdf', methods=['GET'])
def descargar_pdf(doc_id):
    """Descarga el PDF actual (la última versión firmada o el original)."""
    token = request.headers.get('Authorization', '')
    if not decodificar_token(token):
        return jsonify({'ok': False, 'error': 'No autorizado'}), 403

    doc = DocumentoService.obtener_documento(doc_id)
    if not doc:
        return jsonify({'ok': False, 'error': 'Documento no encontrado'}), 404

    ruta = doc.get('ruta_pdf_actual') or doc.get('ruta_pdf_original')
    if not ruta or not os.path.exists(ruta):
        return jsonify({'ok': False, 'error': 'PDF no encontrado'}), 404

    numero = doc.get('numero_accion', 'documento').replace(' ', '_').replace('/', '-')
    nombre_descarga = f'{numero}.pdf'

    return send_file(
        ruta,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=nombre_descarga
    )


@app.route('/api/acciones-personal/<int:doc_id>/excel', methods=['GET'])
def descargar_excel_documento(doc_id):
    """Descarga el Excel guardado del documento."""
    token = request.headers.get('Authorization', '')
    if not decodificar_token(token):
        return jsonify({'ok': False, 'error': 'No autorizado'}), 403

    doc = DocumentoService.obtener_documento(doc_id)
    if not doc:
        return jsonify({'ok': False, 'error': 'Documento no encontrado'}), 404

    ruta = doc.get('ruta_excel')
    if not ruta or not os.path.exists(ruta):
        return jsonify({'ok': False, 'error': 'Excel no encontrado'}), 404

    numero = doc.get('numero_accion', 'documento').replace(' ', '_').replace('/', '-')
    nombre_descarga = f'{numero}.xlsx'

    return send_file(
        ruta,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=nombre_descarga
    )


def _generar_excel_documento(datos, excel_dir, numero_accion):
    """
    Genera el Excel de acción de personal desde datos del formulario.
    Reutiliza la lógica de la ruta /api/generar-accion pero guarda a disco.
    """
    from openpyxl.utils import coordinate_to_tuple
    from datetime import datetime as dt

    def escribir_celda(ws, addr, valor):
        row, col = coordinate_to_tuple(addr)
        for rango in ws.merged_cells.ranges:
            if rango.min_row <= row <= rango.max_row and rango.min_col <= col <= rango.max_col:
                ws.cell(row=rango.min_row, column=rango.min_col, value=valor)
                return
        ws.cell(row=row, column=col, value=valor)

    def parse_fecha(s, como_texto=False):
        if not s:
            return ''
        for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y'):
            try:
                d = dt.strptime(s, fmt)
                return d.strftime('%d/%m/%Y') if como_texto else d
            except ValueError:
                pass
        return s

    carpeta = os.path.join(os.path.dirname(__file__), 'plantillas')
    archivos = glob.glob(os.path.join(carpeta, '*.xlsx'))
    if not archivos:
        raise Exception("No se encontró plantilla Excel en 'plantillas/'.")

    ruta_plantilla = archivos[0]
    wb = openpyxl.load_workbook(ruta_plantilla)
    hoja = wb['ap'] if 'ap' in wb.sheetnames else wb.active

    # Encabezado
    escribir_celda(hoja, 'M3', datos.get('numero_accion', ''))
    escribir_celda(hoja, 'K5', parse_fecha(datos.get('fecha_elaboracion', '')))

    # Funcionario
    escribir_celda(hoja, 'A6', datos.get('apellidos', ''))
    escribir_celda(hoja, 'I6', datos.get('nombres', ''))
    escribir_celda(hoja, 'E11', datos.get('cedula', ''))
    escribir_celda(hoja, 'I11', parse_fecha(datos.get('desde', datos.get('fecha_rige_desde', '')), como_texto=True))
    escribir_celda(hoja, 'M11', parse_fecha(datos.get('hasta', datos.get('fecha_rige_hasta', '')), como_texto=True))

    # Tipo de acción — marcado robusto con normalización de tildes y espacios
    import unicodedata as _ud

    def _norm(t):
        """Normaliza texto: quita tildes, colapsa espacios, pone en mayúsculas."""
        t = str(t or '').strip().upper()
        t = _ud.normalize('NFD', t)
        t = ''.join(c for c in t if _ud.category(c) != 'Mn')
        return ' '.join(t.split())

    MARCAS_ACCION = {
        'INGRESO':               'A14',
        'REINGRESO':             'A15',
        'RESTITUCION':           'A16',
        'REINTEGRO':             'A17',
        'ASCENSO':               'A18',
        'TRASLADO':              'A19',
        'TRASPASO':              'D14',
        'CAMBIO ADMINISTRATIVO': 'D15',
        'INTERCAMBIO VOLUNTARIO':'D16',
        'LICENCIA':              'D17',
        'COMISION DE SERVICIOS': 'D18',
        'SANCIONES':             'D19',
        'INCREMENTO RMU':        'I14',
        'SUBROGACION':           'I15',
        'ENCARGO':               'I16',
        'CESACION DE FUNCIONES': 'I17',
        'DESTITUCION':           'I18',
        'VACACIONES':            'I19',
        'REVISION CLAS. PUESTO': 'L14',
        'OTRO':                  'L15',
    }

    tipo_raw = datos.get('accion_personal') or datos.get('tipo_accion') or ''
    tipo_norm = _norm(tipo_raw)

    celda_marca = MARCAS_ACCION.get(tipo_norm)

    if celda_marca:
        from openpyxl.utils import coordinate_to_tuple as _ctt
        from openpyxl.styles import Font as _Font
        escribir_celda(hoja, celda_marca, 'X')
        _r, _c = _ctt(celda_marca)
        _celda = hoja.cell(row=_r, column=_c)
        _celda.font = _Font(bold=True, name=_celda.font.name, size=_celda.font.size)
    else:
        # Fallback: marcar OTRO y escribir el detalle
        from openpyxl.utils import coordinate_to_tuple as _ctt
        from openpyxl.styles import Font as _Font
        escribir_celda(hoja, 'L15', 'X')
        _r, _c = _ctt('L15')
        _celda = hoja.cell(row=_r, column=_c)
        _celda.font = _Font(bold=True, name=_celda.font.name, size=_celda.font.size)
        if tipo_norm not in ('', 'OTRO'):
            escribir_celda(hoja, 'L16', tipo_raw.upper())

    # Motivación
    escribir_celda(hoja, 'A24', datos.get('motivo_legal', ''))

    # Situación actual
    proc_actual = datos.get('proceso_institucional_actual', '')
    unidad_act = datos.get('unidad', datos.get('unidad_administrativa', ''))
    lugar_act = datos.get('lugar_trabajo_actual', '') or datos.get('ciudad', '')
    denom_act = datos.get('denominacion_actual', '') or datos.get('cargo', '')
    grupo = datos.get('grupo_ocupacional', '')
    partida_act = datos.get('partida_actual', '')
    nivel_gest_act = datos.get('nivel_gestion_actual', '')

    escribir_celda(hoja, 'B28', proc_actual)
    escribir_celda(hoja, 'B30', nivel_gest_act)
    escribir_celda(hoja, 'B32', unidad_act)
    escribir_celda(hoja, 'B34', lugar_act)
    escribir_celda(hoja, 'B36', denom_act)
    escribir_celda(hoja, 'B38', grupo)
    escribir_celda(hoja, 'B44', partida_act)

    # Situación propuesta
    nivel_gest_prop = datos.get('nivel_gestion_propuesta', '') or nivel_gest_act
    proc_prop = datos.get('proceso_institucional_propuesta', '') or proc_actual
    unidad_prop = datos.get('unidad_propuesta', datos.get('unidad_administrativa_propuesta', '')) or unidad_act
    lugar_prop = datos.get('lugar_trabajo_propuesta', '') or lugar_act
    denom_prop = datos.get('denominacion_propuesta', '') or denom_act
    partida_prop = datos.get('partida_propuesta', '') or partida_act

    escribir_celda(hoja, 'J28', proc_prop)
    escribir_celda(hoja, 'J30', nivel_gest_prop)
    escribir_celda(hoja, 'J32', unidad_prop)
    escribir_celda(hoja, 'J34', lugar_prop)
    escribir_celda(hoja, 'J36', denom_prop)
    escribir_celda(hoja, 'J38', grupo)
    escribir_celda(hoja, 'J44', partida_prop)

    # Posesión del puesto
    nombre_posesion = f"{datos.get('apellidos', '')} {datos.get('nombres', '')}".strip()
    if nombre_posesion:
        escribir_celda(hoja, 'C48', nombre_posesion)
    if datos.get('cedula'):
        escribir_celda(hoja, 'N48', datos.get('cedula'))
    escribir_celda(hoja, 'C50', datos.get('ciudad', ''))

    # Responsables de aprobación
    escribir_celda(hoja, 'C61', datos.get('nombre_director_th', ''))
    escribir_celda(hoja, 'C62', datos.get('puesto_director_th', ''))
    escribir_celda(hoja, 'K61', datos.get('nombre_autoridad', ''))
    escribir_celda(hoja, 'K62', datos.get('puesto_autoridad', ''))

    escribir_celda(hoja, 'C61', datos.get('nombre_responsable_th', ''))
    escribir_celda(hoja, 'C62', datos.get('puesto_responsable_th', ''))

    # Aceptación del servidor
    escribir_celda(hoja, 'C74', datos.get('aceptacion_servidor', ''))
    fecha_acep = parse_fecha(
        datos.get('fecha_aceptacion', '') or datos.get('fecha_elaboracion', ''),
        como_texto=True
    )
    if fecha_acep:
        escribir_celda(hoja, 'C75', fecha_acep)

    # Responsables elaboración / revisión / registro
    escribir_celda(hoja, 'C87', datos.get('elaborado_por', ''))
    escribir_celda(hoja, 'C88', datos.get('puesto_elaborado', ''))
    escribir_celda(hoja, 'G87', datos.get('revisado_por', ''))
    escribir_celda(hoja, 'G88', datos.get('puesto_revisado', ''))
    escribir_celda(hoja, 'M87', datos.get('registrado_por', ''))
    escribir_celda(hoja, 'M88', datos.get('puesto_registrado', ''))

    # Guardar en disco
    nombre_seguro = (numero_accion or 'documento').replace(' ', '_').replace('/', '-')
    excel_path = os.path.join(excel_dir, f'{nombre_seguro}.xlsx')
    wb.save(excel_path)

    return excel_path


# =========================
# 🚀 RUN
# =========================
if __name__ == '__main__':
    app.run(port=FLASK_PORT, debug=FLASK_DEBUG)
